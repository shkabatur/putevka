# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
from pprint import pprint
from fpdf import FPDF
import datetime
import json
import re
import sys
from tkinter import messagebox
import logging

logging.basicConfig(level=logging.DEBUG,filename='app.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s')
logging.info("Start logging.....")


sys.stderr = open("errors.txt", "w")
    

def get_age(a,b):
    return str(int((b - a).days / 365))    


def printKind(kind,pdf, pos):
    def print_xy(x,y,text):
        pdf.set_xy(x,y)
        pdf.cell(0,0,text)


    def printInCells(s, x, y, interval):
        for c in s :
            print_xy(x,y,c)
            x += interval

    pdf.add_page()
    pdf.set_font("KEK", size=12)

    for pice in pos:
        name = pice.get("name")
        x1 = pice.get("x1")
        y1 = pice.get("y1")
        x2 = pice.get("x2")
        y2 = pice.get("y2")
        interval = pice.get("interval")
        size = pice.get("size", 12)
        width = pice.get("width")
        length = pice.get("length")

        pdf.set_font("KEK", size=size)

        if interval : # Если это фамилия, имя или отчество
            printInCells(kind[name],x1,y1, interval)
            printInCells(kind[name],x2,y2, interval)
        elif width:  # Если это линия
            if (kind["summa"] and name=="line2") or ( (not kind["summa"]) and name=="line1"):
                pdf.set_line_width(width)
                pdf.line(x1,y1,x1+length,y1)
                pdf.line(x2,y2,x2+length,y2)
        elif name == "summa" and kind["summa"]:
            print_xy(x1,y1,kind["summa"])
            print_xy(x2,y2,kind["summa"])
        elif name == "z" and (not kind["summa"]):
            print_xy(x1,y1,kind["z"])
            print_xy(x2,y2,kind["z"])
        elif name == "ministerstvo1" or name == "ministerstvo2":
            print_xy(x1,y1,kind[name])
        elif name != "z":
            print_xy(x1,y1,kind[name])
            print_xy(x2,y2,kind[name])

def processKinds(smena_no,date_smena,file_from, file_to,s,po):

    DATE = "D"
    ADDRESS = "K"
    NAME = "C"
    RODITEL = "M"
    MINISTERSTVO = "Q"
    SUMMA = "P"

    with open("pos.json") as json_file:
        position = json.load(json_file)["pos"]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_font("KEK", '', 'times-new-roman.ttf', uni=True)

    logging.info(file_from)
    logging.info(file_to)    

    workbook = load_workbook(filename=file_from)
    sheet = workbook.active
    kinds= []

    i = 10
    while sheet[NAME + str(i)].value:
        kind = {}
        kind["smena_no"] = smena_no
        kind["s_d"],kind["s_m"],kind["s_g"] = s
        kind["po_d"],kind["po_m"],kind["po_g"] = po
        
        flp = sheet[NAME + str(i)].value.split()
        if len(flp) == 2:
            kind["last_name"], kind["first_name"] = flp
            kind["patronymic"] = ""
        elif len(flp) == 3:
            kind["last_name"], kind["first_name"], kind["patronymic"] = flp
        else:
            messagebox.showerror("С этой фамилией что-то не так: ",sheet[NAME + str(i)].value)
        
        kind["date"] = sheet[DATE + str(i)].value

        #new 
        parents = sheet[RODITEL + str(i)].value
        if parents:
            parents = re.split("\n|,|  ", sheet[RODITEL + str(i)].value)
            if len(parents) > 1:
                kind["parent1"], kind["parent2"] = parents
            else:
                kind["parent1"] = parents[0]
                kind["parent2"] = ""
        else:
            kind["parent1"] = ""
            kind["parent2"] = ""
        
        ministerstvo = sheet[MINISTERSTVO + str(i)].value
        if ministerstvo:
            if len(ministerstvo) > 50:
                sp = ministerstvo.split(' ')
                kind["ministerstvo1"], kind["ministerstvo2"] = [" ".join(sp[:(len(sp)//2)+2])," ".join(sp[(len(sp)//2)+2:])]
            else:
                kind["ministerstvo1"] = ministerstvo
                kind["ministerstvo2"] = ""
        else:
            kind["ministerstvo1"] = ""
            kind["ministerstvo2"] = ""
        # Адрес
        address = sheet[ADDRESS + str(i)].value
        if address :
            if len(address) > 40:
                sp = address.split(',')
                kind["parent_addr1"],kind["parent_addr2"] = [",".join(sp[:(len(sp)//2)]),",".join(sp[(len(sp)//2):])]
            else:
                kind["parent_addr1"] = address
                kind["parent_addr2"] = ""
        else:
            kind["parent_addr1"] = ""
            kind["parent_addr2"] = ""

        kind["age"] = kind["date"].strftime("%d.%m.%Y") + " (" + get_age(kind["date"],date_smena) + " лет)"
        summa = sheet[SUMMA + str(i)].value
        if summa:
            kind["summa"] = summa
        else:
            kind["summa"] = ""
        kind["z"] = "Ƶ"
        kinds.append(kind)
        i += 1

    for kind in kinds:
        printKind(kind,pdf,position)
    pdf.output(file_to)
    pdf.close()
    messagebox.showinfo("Ура!","Путёвки созданы!")
