# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
from pprint import pprint
from fpdf import FPDF
import datetime
import json
import re
import sys
from tkinter import messagebox
import logging

logging.basicConfig(level=logging.DEBUG,filename='app.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s')
logging.info("Start logging.....")


sys.stderr = open("errors.txt", "w")
    

def getAges(a,b):
    return str(int((b - a).days / 365))    


def printKind(kind,pdf, position,smena_no,s_d,s_m,s_g,po_d,po_m,po_g):
    def print_xy(x,y,text):
        pdf.set_xy(x,y)
        pdf.cell(0,0,text)


    def printInCells(s, x, y):
        INTERVAL = 6.2
        for c in s :
            print_xy(x,y,c)
            x += INTERVAL

    pdf.add_page()
    pdf.set_font("KEK", size=12)
    #==========================================================================
    #========Первая страничка==================================================
    #nomer smeny
    print_xy(position["1s"]["nomer_smeny"]["x"], position["1s"]["nomer_smeny"]["y"], smena_no)

    #S DATE
    print_xy(position["1s"]["s_d"]["x"], position["1s"]["s_d"]["y"], s_d)
    print_xy(position["1s"]["s_m"]["x"], position["1s"]["s_m"]["y"], s_m)
    print_xy(position["1s"]["s_g"]["x"], position["1s"]["s_g"]["y"], s_g)

    #PO DATE
    print_xy(position["1s"]["po_d"]["x"], position["1s"]["po_d"]["y"], po_d)
    print_xy(position["1s"]["po_m"]["x"], position["1s"]["po_m"]["y"], po_m)
    print_xy(position["1s"]["po_g"]["x"], position["1s"]["po_g"]["y"], po_g)

    #VOZRAST
    print_xy(position["1s"]["vozrast"]["x"], position["1s"]["vozrast"]["y"], kind["ages"])

    #FAMILIYA
    printInCells(kind["last_name"], position["1s"]["Familiya"]["x"], position["1s"]["Familiya"]["y"])
    #IMYA
    printInCells(kind["first_name"], position["1s"]["Imya"]["x"], position["1s"]["Imya"]["y"])
    #OTCHESTVO
    printInCells(kind["patronymic"], position["1s"]["Otchestvo"]["x"], position["1s"]["Otchestvo"]["y"])

    #FIO_RODITELYA
    print_xy(position["1s"]["FIO_roditelya"]["x"],position["1s"]["FIO_roditelya"]["y"], kind["parent"][0])
    if len(kind["parent"]) > 1:
        print_xy(position["1s"]["FIO_roditelya"]["x"],position["1s"]["FIO_roditelya"]["y"] + 7, kind["parent"][1])


    #Adres_roditelya
    print_xy(position["1s"]["Adres_roditelya"]["x"],position["1s"]["Adres_roditelya"]["y"], kind["address"][0])
    if len(kind["address"]) > 1:
        print_xy(position["1s"]["Adres_roditelya2"]["x"],position["1s"]["Adres_roditelya2"]["y"], kind["address"][1])

    #Ministerstvo
    print_xy(position["1s"]["ministerstvo"]["x"],position["1s"]["ministerstvo"]["y"], kind["ministerstvo"][0])
    if len(kind["ministerstvo"]) > 1:
        print_xy(position["1s"]["ministerstvo2"]["x"],position["1s"]["ministerstvo2"]["y"], kind["ministerstvo"][1])
    
    #summa or zline
    if kind["summa"]:
        print_xy(position["1s"]["summa"]["x"],position["1s"]["summa"]["y"], kind["summa"])
        x,y = position["1s"]["line2"]["x"],position["1s"]["line2"]["y"]
        width = position["1s"]["line2"]["width"]
        length = position["1s"]["line2"]["length"]
        pdf.set_line_width(width)
        pdf.line(x,y,x+length,y)
    else:
        z_size = position["1s"]["z"]["size"]
        pdf.set_font("KEK", size= z_size)
        print_xy(position["1s"]["z"]["x"],position["1s"]["z"]["y"], "Ƶ")
        x,y = position["1s"]["line"]["x"],position["1s"]["line"]["y"]
        width = position["1s"]["line"]["width"]
        length = position["1s"]["line"]["length"]
        pdf.set_line_width(width)
        pdf.line(x,y,x+length,y)
    pdf.set_font("KEK", size=12)
    #===================Конец первой странички==================================
    #===========================================================================
    
    #==========================================================================
    #========Вторая страничка==================================================
    #nomer smeny
    print_xy(position["2s"]["nomer_smeny"]["x"], position["2s"]["nomer_smeny"]["y"], smena_no)

    #S DATE
    print_xy(position["2s"]["s_d"]["x"], position["2s"]["s_d"]["y"], s_d)
    print_xy(position["2s"]["s_m"]["x"], position["2s"]["s_m"]["y"], s_m)
    print_xy(position["2s"]["s_g"]["x"], position["2s"]["s_g"]["y"], s_g)

    #PO DATE
    print_xy(position["2s"]["po_d"]["x"], position["2s"]["po_d"]["y"], po_d)
    print_xy(position["2s"]["po_m"]["x"], position["2s"]["po_m"]["y"], po_m)
    print_xy(position["2s"]["po_g"]["x"], position["2s"]["po_g"]["y"], po_g)

    #VOZRAST
    print_xy(position["2s"]["vozrast"]["x"], position["2s"]["vozrast"]["y"], kind["ages"])

    #FAMILIYA
    printInCells(kind["last_name"], position["2s"]["Familiya"]["x"], position["2s"]["Familiya"]["y"])
    #IMYA
    printInCells(kind["first_name"], position["2s"]["Imya"]["x"], position["2s"]["Imya"]["y"])
    #OTCHESTVO
    printInCells(kind["patronymic"], position["2s"]["Otchestvo"]["x"], position["2s"]["Otchestvo"]["y"])

    #FIO_RODITELYA
    print_xy(position["2s"]["FIO_roditelya"]["x"],position["2s"]["FIO_roditelya"]["y"], kind["parent"][0])
    if len(kind["parent"]) > 1:
        print_xy(position["2s"]["FIO_roditelya"]["x"],position["2s"]["FIO_roditelya"]["y"] + 7, kind["parent"][1])


    #Adres_roditelya
    print_xy(position["2s"]["Adres_roditelya"]["x"],position["2s"]["Adres_roditelya"]["y"], kind["address"][0])
    if len(kind["address"]) > 1:
        print_xy(position["2s"]["Adres_roditelya2"]["x"],position["2s"]["Adres_roditelya2"]["y"], kind["address"][1])

    #summa or zline
    if kind["summa"]:
        print_xy(position["2s"]["summa"]["x"],position["2s"]["summa"]["y"], kind["summa"])
        x,y = position["2s"]["line2"]["x"],position["2s"]["line2"]["y"]
        width = position["2s"]["line2"]["width"]
        length = position["2s"]["line2"]["length"]
        pdf.set_line_width(width)
        pdf.line(x,y,x+length,y)
    else:
        z_size = position["2s"]["z"]["size"]
        pdf.set_font("KEK", size= z_size)
        print_xy(position["2s"]["z"]["x"],position["2s"]["z"]["y"], "Ƶ")
        x,y = position["2s"]["line"]["x"],position["2s"]["line"]["y"]
        width = position["2s"]["line"]["width"]
        length = position["2s"]["line"]["length"]
        pdf.set_line_width(width)
        pdf.line(x,y,x+length,y)

    #===================Конец второй странички==================================
    #===========================================================================

def processKinds(smena_no,date_smena,file_from, file_to,s,po):

    DATE = "D"
    ADDRESS = "K"
    NAME = "C"
    RODITEL = "M"
    MINISTERSTVO = "Q"
    SUMMA = "P"

    with open("position.json") as json_file:
        position = json.load(json_file)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_font("KEK", '', 'times-new-roman.ttf', uni=True)

    logging.info(file_from)
    logging.info(file_to)    
    s_d, s_m, s_g = s
    po_d, po_m, po_g = po

    workbook = load_workbook(filename=file_from)
    sheet = workbook.active
    kinds= []

    i = 10
    while sheet[NAME + str(i)].value:
        kind = {}
        flp = sheet[NAME + str(i)].value.split()
        if len(flp) == 2:
            kind["last_name"], kind["first_name"] = flp
            kind["patronymic"] = ""
        else:
            kind["last_name"], kind["first_name"], kind["patronymic"] = flp
        kind["date"] = sheet[DATE + str(i)].value
        kind["parent"] = re.split("\n|,|  ", sheet[RODITEL + str(i)].value)
        
        ministerstvo = sheet[MINISTERSTVO + str(i)].value
        if len(ministerstvo) > 50:
            sp = ministerstvo.split(' ')
            kind["ministerstvo"] = [" ".join(sp[:(len(sp)//2)+2])," ".join(sp[(len(sp)//2)+2:])]
        else:
            kind["ministerstvo"] = [ministerstvo]
        
        address = sheet[ADDRESS + str(i)].value
        if address :
            if len(address) > 40:
                sp = address.split(',')
                kind["address"] = [",".join(sp[:(len(sp)//2)]),",".join(sp[(len(sp)//2):])]
            else:
                kind["address"] = [address]
        else:
            kind["address"] = [""]
        kind["ages"] = kind["date"].strftime("%d.%m.%Y") + " (" + getAges(kind["date"],date_smena) + " лет)"
        summa = sheet[SUMMA + str(i)].value
        if summa:
            kind["summa"] = summa
        else:
            kind["summa"] = ""
        kinds.append(kind)
        i += 1

    for kind in kinds:
        printKind(kind,pdf,position,smena_no,s_d,s_m,s_g,po_d,po_m,po_g)
    pdf.output(file_to)
    pdf.close()
    messagebox.showinfo("Ура!","Путёвки созданы!")
