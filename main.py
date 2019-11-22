# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
from pprint import pprint
from fpdf import FPDF
import datetime
import json
import re
import sys

with open("position.json") as json_file:
    position = json.load(json_file)

log = open("log.txt", "w")
sys.stderr = open("errors.txt", "w")

pdf = FPDF(orientation="L", unit="mm", format="A4")
pdf.add_font("KEK", '', 'times-new-roman.ttf', uni=True)

DATE = "D"
ADDRESS = "K"
NAME = "C"
RODITEL = "M"
MINISTERSTVO = "Q"
print(sys.argv[1:], file=log)
file_from, file_to, SMENA, date_smena, s, po = sys.argv[1:]
file_to = file_to + datetime.datetime.now().strftime("%H.%M.%S") + ".pdf"


ss = s.split('-')
print("ss:",ss,file=log)
if len(ss) == 4:
    s_den1,s_den2, s_mesyac, s_god = ss
    s_den = "-".join([s_den1, s_den2])
else:
    s_den, s_mesyac, s_god = ss

ppo = po.split("-")
print("ppo = ", ppo, file=log)
if len(ppo) == 4:
    po_den1,po_den2, po_mesyac, po_god = ppo
    po_den = "-".join([po_den1, po_den2])
else:
    po_den, po_mesyac, po_god = ppo

y,m,d = date_smena.split('.')
date_smena = datetime.datetime(int(y), int(m), int(d))



workbook = load_workbook(filename=file_from)
sheet = workbook.active
kinds= []

def getAges(a):
    return str(int((date_smena - a).days / 365))    


def print_xy(x,y,text):
    pdf.set_xy(x,y)
    pdf.cell(0,0,text)

def printInCells(s, x, y):
    INTERVAL = 6.2
    for c in s :
        print_xy(x,y,c)
        x += INTERVAL


def printKind(kind):
    pdf.add_page()
    pdf.set_font("KEK", size=12)
    #==========================================================================
    #========Первая страничка==================================================
    #nomer smeny
    print_xy(position["1s"]["nomer_smeny"]["x"], position["1s"]["nomer_smeny"]["y"], SMENA)

    #S DATE
    print_xy(position["1s"]["s_den"]["x"], position["1s"]["s_den"]["y"], s_den)
    print_xy(position["1s"]["s_mesyac"]["x"], position["1s"]["s_mesyac"]["y"], s_mesyac)
    print_xy(position["1s"]["s_god"]["x"], position["1s"]["s_god"]["y"], s_god)

    #PO DATE
    print_xy(position["1s"]["po_den"]["x"], position["1s"]["po_den"]["y"], po_den)
    print_xy(position["1s"]["po_mesyac"]["x"], position["1s"]["po_mesyac"]["y"], po_mesyac)
    print_xy(position["1s"]["po_god"]["x"], position["1s"]["po_god"]["y"], po_god)

    #VOZRAST
    print_xy(position["1s"]["vozrast"]["x"], position["1s"]["vozrast"]["y"], kind["ages"])

    #FAMILIYA
    printInCells(kind["first_name"], position["1s"]["Familiya"]["x"], position["1s"]["Familiya"]["y"])
    #IMYA
    printInCells(kind["last_name"], position["1s"]["Imya"]["x"], position["1s"]["Imya"]["y"])
    #OTCHESTVO
    printInCells(kind["patronymic"], position["1s"]["Otchestvo"]["x"], position["1s"]["Otchestvo"]["y"])

    #FIO_RODITELYA
    print_xy(position["1s"]["FIO_roditelya"]["x"],position["1s"]["FIO_roditelya"]["y"], kind["parent"][0])
    if len(kind["parent"]) > 1:
        print_xy(position["1s"]["FIO_roditelya"]["x"],position["1s"]["FIO_roditelya"]["y"] + 7, kind["parent"][1])


    #Adres_roditelya
    print_xy(position["1s"]["Adres_roditelya"]["x"],position["1s"]["Adres_roditelya"]["y"], kind["address"][0])
    if len(kind["address"]) > 1:
        print_xy(position["1s"]["Adres_roditelya2"]["x"],position["1s"]["Adres_roditelya2"]["y"], kind["address"][1])

    #Ministerstvo
    print_xy(position["1s"]["ministerstvo"]["x"],position["1s"]["ministerstvo"]["y"], kind["ministerstvo"][0])
    if len(kind["ministerstvo"]) > 1:
        print_xy(position["1s"]["ministerstvo2"]["x"],position["1s"]["ministerstvo2"]["y"], kind["ministerstvo"][1])
    
    #LINE
    #x = position["1s"]["line"]["x"]
    #y = position["1s"]["line"]["y"]
    #pdf.set_line_width(2)
    #pdf.line(x,y,x+35,y)
    #===================Конец первой странички==================================
    #===========================================================================

    #==========================================================================
    #========Вторая страничка==================================================
    #nomer smeny
    print_xy(position["2s"]["nomer_smeny"]["x"], position["2s"]["nomer_smeny"]["y"], SMENA)

    #S DATE
    print_xy(position["2s"]["s_den"]["x"], position["2s"]["s_den"]["y"], s_den)
    print_xy(position["2s"]["s_mesyac"]["x"], position["2s"]["s_mesyac"]["y"], s_mesyac)
    print_xy(position["2s"]["s_god"]["x"], position["2s"]["s_god"]["y"], s_god)

    #PO DATE
    print_xy(position["2s"]["po_den"]["x"], position["2s"]["po_den"]["y"], po_den)
    print_xy(position["2s"]["po_mesyac"]["x"], position["2s"]["po_mesyac"]["y"], po_mesyac)
    print_xy(position["2s"]["po_god"]["x"], position["2s"]["po_god"]["y"], po_god)

    #VOZRAST
    print_xy(position["2s"]["vozrast"]["x"], position["2s"]["vozrast"]["y"], kind["ages"])

    #FAMILIYA
    printInCells(kind["first_name"], position["2s"]["Familiya"]["x"], position["2s"]["Familiya"]["y"])
    #IMYA
    printInCells(kind["last_name"], position["2s"]["Imya"]["x"], position["2s"]["Imya"]["y"])
    #OTCHESTVO
    printInCells(kind["patronymic"], position["2s"]["Otchestvo"]["x"], position["2s"]["Otchestvo"]["y"])

    #FIO_RODITELYA
    print_xy(position["2s"]["FIO_roditelya"]["x"],position["2s"]["FIO_roditelya"]["y"], kind["parent"][0])
    if len(kind["parent"]) > 1:
        print_xy(position["2s"]["FIO_roditelya"]["x"],position["2s"]["FIO_roditelya"]["y"] + 7, kind["parent"][1])


    #Adres_roditelya
    print_xy(position["2s"]["Adres_roditelya"]["x"],position["2s"]["Adres_roditelya"]["y"], kind["address"][0])
    if len(kind["address"]) > 1:
        print_xy(position["2s"]["Adres_roditelya2"]["x"],position["2s"]["Adres_roditelya2"]["y"], kind["address"][1])


    #LINE
    #x = position["2s"]["line"]["x"]
    #y = position["2s"]["line"]["y"]
    #pdf.set_line_width(2)
    #pdf.line(x,y,x+35,y)
    #===================Конец второйстранички==================================
    #===========================================================================


i = 10
while sheet[NAME + str(i)].value:
    kind = {}
    flp = sheet[NAME + str(i)].value.split()
    if len(flp) == 2:
         kind["first_name"], kind["last_name"] = flp
         kind["patronymic"] = ""
    else:
        kind["first_name"], kind["last_name"], kind["patronymic"] = flp
    kind["date"] = sheet[DATE + str(i)].value
    kind["parent"] = re.split("\n|,|  ", sheet[RODITEL + str(i)].value)
    
    ministerstvo = sheet[MINISTERSTVO + str(i)].value
    if len(ministerstvo) > 50:
        sp = ministerstvo.split(' ')
        kind["ministerstvo"] = [" ".join(sp[:(len(sp)//2)+2])," ".join(sp[(len(sp)//2)+2:])]
    else:
        kind["ministerstvo"] = [ministerstvo]

    address = sheet[ADDRESS + str(i)].value
    if address :
        if len(address) > 40:
            sp = address.split(',')
            kind["address"] = [",".join(sp[:(len(sp)//2)]),",".join(sp[(len(sp)//2):])]
        else:
            kind["address"] = [address]
    else:
        kind["address"] = [""]
    kind["ages"] = kind["date"].strftime("%d.%m.%Y") + " (" + getAges(kind["date"]) + " лет)"
    kinds.append(kind)
    i += 1

def main():
    for kind in kinds:
        printKind(kind)
    pdf.output(file_to)



if __name__ == "__main__":
    main()
